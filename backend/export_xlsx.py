"""Server-side polished .xlsx generation -- same shape/style as the original
09_inactive_accounts/decisions_to_xlsx.py, but built from live SQLite state
instead of a CSV a reviewer exported by hand. Returns bytes (an in-memory
workbook), never touches disk -- the API layer streams it straight back.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from db import get_connection

INACTIVE_DECISION_ORDER = ["Keep", "Archive", "Merge", "Delete", "Undecided"]
DUPLICATE_DECISION_ORDER = ["Merge", "Not a duplicate", "Keep separate", "Delete", "Undecided"]


def autosize(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str))) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_workbook(df: pd.DataFrame, decision_order: list[str]) -> bytes:
    summary = (
        df["decision_label"].value_counts().reindex(decision_order).fillna(0).astype(int)
        .rename_axis("decision").reset_index(name="count")
    )
    summary.loc[len(summary)] = ["Total", len(df)]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        autosize(writer.sheets["Summary"], summary)

        for decision in decision_order:
            sheet_df = df[df["decision_label"] == decision].drop(columns=["decision_label"])
            if sheet_df.empty:
                continue
            sheet_df.to_excel(writer, sheet_name=decision, index=False)
            autosize(writer.sheets[decision], sheet_df)
    return buf.getvalue()


def build_inactive_xlsx() -> bytes:
    conn = get_connection()
    rows = conn.execute(
        """SELECT c.accountid, c.name, c.dormancy_tier,
                  COALESCE(d.decision, '') AS decision,
                  COALESCE(d.merge_target_accountid, '') AS merge_target_accountid,
                  COALESCE(d.merge_target_name, '') AS merge_target_name,
                  COALESCE(d.note, '') AS note,
                  COALESCE(d.reviewer, '') AS reviewer,
                  COALESCE(d.decided_at, '') AS decided_at
           FROM inactive_candidates c
           LEFT JOIN inactive_decisions d ON d.accountid = c.accountid"""
    ).fetchall()
    conn.close()

    df = pd.DataFrame([dict(r) for r in rows])
    df["decision_label"] = df["decision"].replace("", "Undecided")
    return _write_workbook(df, INACTIVE_DECISION_ORDER)


def build_duplicates_xlsx() -> bytes:
    conn = get_connection()
    # Include historical (already-merged) members only when they carry a
    # decision (i.e. marked for deletion) -- pending members are always
    # included. A plain member with no decision and already merged away is
    # historical context, nothing to act on, so it's excluded.
    rows = conn.execute(
        """SELECT m.cluster_id, cl.signals, cl.confidence, m.accountid, m.name,
                  m.is_already_merged_away, m.existing_masterid_name,
                  pc.accountid AS confirmed_primary_accountid,
                  COALESCE(d.decision, '') AS decision,
                  COALESCE(d.note, '') AS note,
                  COALESCE(d.reviewer, '') AS reviewer,
                  COALESCE(d.decided_at, '') AS decided_at
           FROM duplicate_cluster_members m
           JOIN duplicate_clusters cl ON cl.cluster_id = m.cluster_id
           LEFT JOIN duplicate_decisions d ON d.accountid = m.accountid
           LEFT JOIN duplicate_primary_choices pc ON pc.cluster_id = m.cluster_id
           WHERE m.is_already_merged_away = 0
              OR COALESCE(d.decision, '') != ''"""
    ).fetchall()
    conn.close()

    records = []
    for r in rows:
        r = dict(r)
        is_primary = bool(r["confirmed_primary_accountid"]) and r["confirmed_primary_accountid"] == r["accountid"]
        records.append({
            "cluster_id": r["cluster_id"],
            "signals": r["signals"],
            "confidence": r["confidence"],
            "accountid": r["accountid"],
            "name": r["name"],
            "is_already_merged_away": int(r["is_already_merged_away"]),
            "merged_into": r["existing_masterid_name"] if r["is_already_merged_away"] else "",
            "is_primary": int(is_primary),
            "decision": "" if is_primary else r["decision"],
            "note": r["note"],
            "reviewer": r["reviewer"],
            "decided_at": r["decided_at"],
        })

    df = pd.DataFrame(records)
    df["decision_label"] = df["decision"].replace("", "Undecided")
    return _write_workbook(df, DUPLICATE_DECISION_ORDER)


def build_contact_duplicates_xlsx() -> bytes:
    conn = get_connection()
    rows = conn.execute(
        """SELECT m.cluster_id, cl.signals, cl.confidence, m.contactid, m.fullname,
                  m.emailaddress1, m.jobtitle, m.parent_account_name,
                  m.is_already_merged_away, m.existing_masterid_name,
                  pc.contactid AS confirmed_primary_contactid,
                  COALESCE(d.decision, '') AS decision,
                  COALESCE(d.note, '') AS note,
                  COALESCE(d.reviewer, '') AS reviewer,
                  COALESCE(d.decided_at, '') AS decided_at
           FROM duplicate_contact_cluster_members m
           JOIN duplicate_contact_clusters cl ON cl.cluster_id = m.cluster_id
           LEFT JOIN duplicate_contact_decisions d ON d.contactid = m.contactid
           LEFT JOIN duplicate_contact_primary_choices pc ON pc.cluster_id = m.cluster_id
           WHERE m.is_already_merged_away = 0
              OR COALESCE(d.decision, '') != ''"""
    ).fetchall()
    conn.close()

    records = []
    for r in rows:
        r = dict(r)
        is_primary = bool(r["confirmed_primary_contactid"]) and r["confirmed_primary_contactid"] == r["contactid"]
        records.append({
            "cluster_id": r["cluster_id"],
            "signals": r["signals"],
            "confidence": r["confidence"],
            "contactid": r["contactid"],
            "fullname": r["fullname"],
            "emailaddress1": r["emailaddress1"],
            "jobtitle": r["jobtitle"],
            "parent_account_name": r["parent_account_name"],
            "is_already_merged_away": int(r["is_already_merged_away"]),
            "merged_into": r["existing_masterid_name"] if r["is_already_merged_away"] else "",
            "is_primary": int(is_primary),
            "decision": "" if is_primary else r["decision"],
            "note": r["note"],
            "reviewer": r["reviewer"],
            "decided_at": r["decided_at"],
        })

    df = pd.DataFrame(records)
    df["decision_label"] = df["decision"].replace("", "Undecided")
    return _write_workbook(df, DUPLICATE_DECISION_ORDER)
